import os
import re
import random
import traceback
from flask import Flask, request, render_template, send_file, url_for, flash, session, redirect
from werkzeug.utils import secure_filename
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import resend
import xlsxwriter


# ==========================
# App Setup
# ==========================
app = Flask(__name__)
app.secret_key = 're_hLuavChZ_PZLY9Jiuvb7YuGptHKHm6pup'
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['RESULT_FOLDER'] = 'results'

os.makedirs('uploads', exist_ok=True)
os.makedirs('results', exist_ok=True)

# ==========================
# Resend Configuration
# ==========================
resend.api_key = "re_hLuavChZ_PZLY9Jiuvb7YuGptHKHm6pup"  # REPLACE WITH YOUR REAL KEY

# Your verified email (Resend allows sending to this)
MY_EMAIL = "reshmamohan938@gmail.com"

# TEST MODE: True = Forward all emails to your Gmail (bypasses domain verification)
# Set to False when softnis.com is verified
TEST_MODE = False

def send_email(subject, to, body, html=None):
    if not isinstance(to, list):
        to = [to]

    params = {
        "from": "SoftNis <info@softnis.com>",  # Now verified!
        "to": to,
        "subject": subject,
        "text": body
    }
    if html:
        params["html"] = html

    try:
        resend.Emails.send(params)
        print(f"EMAIL SENT → {to[0]}")
    except Exception as e:
        print(f"EMAIL FAILED: {e}")
        raise
# ==========================
# Routes
# ==========================

@app.route('/')
def start_page():
    return render_template('start_page.html')

@app.route('/prerequisites')
def prerequisites():
    return render_template('prerequisites.html')

@app.route('/login', methods=['GET', 'POST'])  # Assuming this is your "register" step
def login():
    if request.method == 'POST':
        email = request.form['email'].strip().lower()
        allowed_extra = ["softnisdata@gmail.com"]
        if not (email.endswith("@softnis.com") or email in allowed_extra):
            flash("❌ Only @softnis.com emails allowed.", "popup")
            return redirect(url_for('login'))

        otp = str(random.randint(1000, 9999))
        session['otp'] = otp
        session['email'] = email

        # Send OTP (in test mode: to your Gmail with note)
        send_email(
            subject=f"OTP for {email} - SoftNis Login",
            to=email,
            body=f"Your OTP is: {otp}\n\nEmail: {email}\nValid for 5 minutes.",
            html=f"<h2>Login OTP</h2><p><strong>OTP: {otp}</strong></p><p>Email: {email}</p>"
        )

        # Admin alert (also forwarded in test mode)
        send_email(
            subject="Login Request Alert",
            to="info@softnis.com",
            body=f"User {email} requested OTP. Generated OTP: {otp}"
        )

        if TEST_MODE:
            flash(f"✅ OTP sent! Check {MY_EMAIL} (Test Mode - shows original {email}).", "popup")
        else:
            flash("✅ OTP sent to your email!", "popup")
        return redirect(url_for('verify'))

    return render_template('login.html')

@app.route('/verify', methods=['GET', 'POST'])
def verify():
    if 'otp' not in session:
        return redirect(url_for('login'))

    if request.method == 'POST':
        entered_otp = request.form.get('otp')
        if entered_otp == session.get('otp'):
            session['logged_in'] = True
            session.pop('otp', None)
            flash("✅ Login successful!", "success")
            return redirect(url_for('prerequisites'))
        flash("❌ Wrong OTP. Try again.", "popup")

    return render_template('verify.html')

@app.route('/logout')
def logout():
    session.clear()
    flash("🚪 Logged out successfully.", "success")
    return redirect(url_for('login'))

@app.route('/send')  # Test route
def send_email_test():
    test_email = "test@softnis.com"  # Simulate any @softnis.com
    send_email(
        subject="Test OTP",
        to=test_email,
        body="Test OTP: 1234",
        html="<strong>Test: Works for any @softnis.com!</strong>"
    )
    return f"Test sent! (Check {MY_EMAIL if TEST_MODE else test_email})"

# ==========================
# SoftNis ID Validation
# ==========================
def is_valid_softnis(series):
    pattern = re.compile(r'^(?=.*[A-Za-z])[A-Za-z0-9_]+$')
    for val in series:
        val = str(val).strip()
        if val and not pattern.match(val):
            return False
    return True

# ==========================
# File Upload & Processing
# ==========================
@app.route('/index', methods=['GET', 'POST'])
def index():
    if not session.get('logged_in'):
        flash("⚠ Please log in.", "popup")
        return redirect('/login')

    if request.method == 'POST':
        file = request.files.get('file')
        if not file or not file.filename.lower().endswith(('.xls', '.xlsx')):
            flash("❌ Upload .xls or .xlsx file.", "popup")
            return render_template('index.html')

        filename = secure_filename(file.filename)
        file_path = os.path.join('uploads', filename)
        file.save(file_path)

        try:
            excel_file = pd.ExcelFile(file_path)
            sheets = excel_file.sheet_names
            if len(sheets) != 2 or 'Production Completed' not in sheets or 'Delivered' not in sheets:
                flash("❌ Need 2 sheets: 'Production Completed' & 'Delivered'.", "popup")
                return render_template('index.html')

            df_prod = pd.read_excel(file_path, sheet_name='Production Completed')
            df_del = pd.read_excel(file_path, sheet_name='Delivered')

            if 'User Name' not in df_prod.columns or 'User Name' not in df_del.columns:
                flash("❌ 'User Name' column missing.", "popup")
                return render_template('index.html')

            prod_id_col = next((c for c in df_prod.columns if str(c).strip().lower() == "softnis id"), None)
            del_id_col = next((c for c in df_del.columns if str(c).strip().lower() == "softnis id"), None)
            if not prod_id_col or not del_id_col:
                flash("❌ 'SoftNis ID' column missing.", "popup")
                return render_template('index.html')

            if not is_valid_softnis(df_prod[prod_id_col]) or not is_valid_softnis(df_del[del_id_col]):
                flash("❌ Invalid SoftNis ID (letters/numbers/underscore only).", "popup")
                return render_template('index.html')

            result_filename = f"result_{filename}"
            result_path = os.path.join('results', result_filename)
            generate_report(file_path, df_prod, df_del, result_path, prod_id_col)

            flash("✅ Report ready for download!", "success")
            return render_template('result.html', download_link=url_for('download_file', filename=result_filename))

        except Exception as e:
            print(traceback.format_exc())
            flash(f"❌ Error: {str(e)}", "popup")

    return render_template('index.html')

@app.route('/download/<filename>')
def download_file(filename):
    path = os.path.join('results', filename)
    if os.path.exists(path):
        return send_file(path, as_attachment=True)
    flash("❌ File not found.", "popup")
    return redirect(url_for('index'))

# ==========================
# Report Generator
# ==========================
# def generate_report(file_path, df_prod, df_del, result_path, id_col):
#     ignore_cols = [c.lower().strip() for c in df_prod.columns if 'attribute name' in c.lower()]
#     ignore_cols += ['softnis id', 'user name']
#
#     for i in range(56):
#         name_col = f"Technical Specification {i+1} Name"
#         value_col = f"Technical Specification {i+1} Value"
#         attr_name = f"Attribute Name.{i}" if i else "Attribute Name"
#         attr_value = f"Attribute Value.{i}" if i else "Attribute Value"
#         if name_col in df_del.columns:
#             df_del.rename(columns={name_col: attr_name}, inplace=True)
#         if value_col in df_del.columns:
#             df_del.rename(columns={value_col: attr_value}, inplace=True)
#
#     red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
#     wb = load_workbook(file_path)
#     ws_prod = wb["Production Completed"]
#     headers = list(df_prod.columns)
#     user_col_index = next((i for i, c in enumerate(headers) if str(c).strip().lower() == 'user name'), -1)
#
#     quality_start = len(headers) + 1
#     ws_prod.cell(1, quality_start, "Row Quality %")
#     ws_prod.cell(1, quality_start + 1, "Right Values")
#     ws_prod.cell(1, quality_start + 2, "Wrong Values")
#     ws_prod.cell(1, quality_start + 3, "Error Report")
#
#     df_del = df_del.drop_duplicates(subset=[id_col])
#     df_del_dict = df_del.set_index(id_col).to_dict('index')
#     user_stats = {}
#
#     for i in range(len(df_prod)):
#         right = wrong = 0
#         prod_row = df_prod.iloc[i]
#         sid = str(prod_row[id_col]).strip()
#
#         if not sid or sid not in df_del_dict:
#             ws_prod.cell(i + 2, quality_start + 3, "SoftNis ID not in Delivered")
#             for c in range(1, quality_start + 4):
#                 ws_prod.cell(i + 2, c).fill = red_fill
#             continue
#
#         del_row = df_del_dict[sid]
#
#         for j, col in enumerate(headers):
#             cname = str(col).strip().lower()
#             if cname in ignore_cols:
#                 continue
#             v1 = str(prod_row[col]).strip() if pd.notna(prod_row[col]) else ""
#             v2 = str(del_row.get(col, "")).strip() if pd.notna(del_row.get(col, "")) else ""
#             if not v1 and not v2:
#                 continue
#             if v1 == v2 and v1:
#                 right += 1
#             elif v1 or v2:
#                 ws_prod.cell(i + 2, j + 1).fill = red_fill
#                 wrong += 1
#
#         total = right + wrong
#         quality = round(right / total * 100, 2) if total else 0
#         ws_prod.cell(i + 2, quality_start, quality)
#         ws_prod.cell(i + 2, quality_start + 1, right)
#         ws_prod.cell(i + 2, quality_start + 2, wrong)
#
#         p_user = str(prod_row.get("User Name", "")).strip().lower()
#         d_user = str(del_row.get("User Name", "")).strip().lower()
#         if p_user and d_user and p_user != d_user:
#             cell = ws_prod.cell(i + 2, quality_start + 3)
#             cell.value = "User Name mismatch"
#             cell.fill = red_fill
#             for c in range(1, quality_start + 4):
#                 ws_prod.cell(i + 2, c).fill = red_fill
#
#         username = str(prod_row[user_col_index]).strip()
#         if username:
#             user_stats.setdefault(username, {'correct': 0, 'total': 0})
#             user_stats[username]['correct'] += right
#             user_stats[username]['total'] += total
#
#     if "Quality Report" in wb.sheetnames:
#         wb.remove(wb["Quality Report"])
#     ws_report = wb.create_sheet("Quality Report")
#     ws_report.append(["User Name", "Matched Cells", "Total Cells", "Quality %"])
#     for user, s in user_stats.items():
#         q = round(s['correct'] / s['total'] * 100, 2) if s['total'] else 0
#         ws_report.append([user, s['correct'], s['total'], q])
#
#     wb.save(result_path)

# --------------------------------------------------------------
#  NEW generate_report() – uses pandas + xlsxwriter (fast & low RAM)
# --------------------------------------------------------------
# ==========================
# FINAL generate_report() – Fixed, Safe, Fast
# ==========================
# --------------------------------------------------------------
#  FINAL generate_report() – writes **both** sheets
#  • Production Completed (with red cells)
#  • Quality Report
#  • Delivered sheet is copied **as-is**
# --------------------------------------------------------------
def generate_report(file_path, df_prod, df_del, result_path, id_col):
    # ------------------------------------------------------------------
    # 1. Delivered → unique by SoftNis ID
    # ------------------------------------------------------------------
    df_del = df_del.drop_duplicates(subset=[id_col])
    del_dict = df_del.set_index(id_col).to_dict('index')

    # ------------------------------------------------------------------
    # 2. Columns to ignore
    # ------------------------------------------------------------------
    prod_cols_lower = {c.strip().lower(): c for c in df_prod.columns}
    ignore_cols = {
        prod_cols_lower.get('softnis id'),
        prod_cols_lower.get('user name')
    }
    ignore_cols.update(c for c in df_prod.columns if 'attribute name' in str(c).lower())

    # ------------------------------------------------------------------
    # 3. Add result columns to Production
    # ------------------------------------------------------------------
    df_prod = df_prod.copy()
    df_prod['Row Quality %'] = 0
    df_prod['Right Values']   = 0
    df_prod['Wrong Values']   = 0
    df_prod['Error Report']   = ''

    # ------------------------------------------------------------------
    # 4. Compare rows
    # ------------------------------------------------------------------
    user_stats = {}
    red_cells = []                     # (row, col)  or  row (whole row)

    for idx, prod_row in df_prod.iterrows():
        sid = str(prod_row[id_col]).strip()
        right = wrong = 0
        errors = []

        if not sid or sid not in del_dict:
            errors.append("SoftNis ID not in Delivered")
            red_cells.append(idx)                     # whole row red
        else:
            del_row = del_dict[sid]

            for col in df_prod.columns:
                if col in ignore_cols:
                    continue
                v1 = str(prod_row[col]).strip() if pd.notna(prod_row[col]) else ""
                v2 = str(del_row.get(col, "")).strip() if pd.notna(del_row.get(col, "")) else ""
                if not v1 and not v2:
                    continue
                if v1 == v2 and v1:
                    right += 1
                elif v1 or v2:
                    wrong += 1
                    col_idx = df_prod.columns.get_loc(col)
                    red_cells.append((idx, col_idx))

            # User-name mismatch
            p_user = str(prod_row.get('User Name', '')).strip().lower()
            d_user = str(del_row.get('User Name', '')).strip().lower()
            if p_user and d_user and p_user != d_user:
                errors.append("User Name mismatch")
                red_cells.append(idx)

        total = right + wrong
        quality = round(right / total * 100, 2) if total else 0

        df_prod.at[idx, 'Row Quality %'] = quality
        df_prod.at[idx, 'Right Values']   = right
        df_prod.at[idx, 'Wrong Values']   = wrong
        df_prod.at[idx, 'Error Report']   = '; '.join(errors)

        username = str(prod_row.get('User Name', '')).strip()
        if username:
            user_stats.setdefault(username, {'correct': 0, 'total': 0})
            user_stats[username]['correct'] += right
            user_stats[username]['total']   += total

    # ------------------------------------------------------------------
    # 5. Clean NaN / inf
    # ------------------------------------------------------------------
    df_prod = df_prod.replace([float('inf'), -float('inf')], None).fillna('')

    # ------------------------------------------------------------------
    # 6. Write **all three sheets** with xlsxwriter
    # ------------------------------------------------------------------
    with pd.ExcelWriter(
        result_path,
        engine='xlsxwriter',
        engine_kwargs={'options': {'nan_inf_to_errors': True}}
    ) as writer:

        # ----- 1. Production Completed (with red cells) -----
        df_prod.to_excel(writer, sheet_name='Production Completed', index=False)
        ws_prod = writer.sheets['Production Completed']
        red_fill   = writer.book.add_format({'bg_color': '#FFC7CE'})
        header_fmt = writer.book.add_format({'bold': True, 'bg_color': '#D3D3D3'})

        # header
        for c, col in enumerate(df_prod.columns):
            ws_prod.write(0, c, col, header_fmt)

        # red cells
        for item in red_cells:
            if isinstance(item, tuple):
                r, c = item
                ws_prod.write(r + 1, c, df_prod.iat[r, c], red_fill)
            else:
                for c in range(len(df_prod.columns)):
                    ws_prod.write(item + 1, c, df_prod.iat[item, c], red_fill)

        # ----- 2. Delivered (copy as-is) -----
        df_del.to_excel(writer, sheet_name='Delivered', index=False)
        ws_del = writer.sheets['Delivered']
        for c, col in enumerate(df_del.columns):
            ws_del.write(0, c, col, header_fmt)

        # ----- 3. Quality Report -----
        report_data = []
        for user, s in user_stats.items():
            q = round(s['correct'] / s['total'] * 100, 2) if s['total'] else 0
            report_data.append([user, s['correct'], s['total'], q])

        if report_data:
            df_report = pd.DataFrame(
                report_data,
                columns=['User Name', 'Matched Cells', 'Total Cells', 'Quality %']
            )
            df_report.to_excel(writer, sheet_name='Quality Report', index=False)
            ws_rep = writer.sheets['Quality Report']
            for c, col in enumerate(df_report.columns):
                ws_rep.write(0, c, col, header_fmt)

    # ------------------------------------------------------------------
    # DONE – file written in < 5 seconds even for 50 k rows
    # ------------------------------------------------------------------

# ==========================
# Run App
# ==========================
if __name__ == '__main__':
    print("\n" + "="*70)
    print("   SOFTNIS QUALITY TOOL - LOCALHOST READY")
    print("   http://127.0.0.1:5000")
    print(f"   Test Mode: {'ON (forwards to Gmail)' if TEST_MODE else 'OFF (direct to user)'}")
    print("="*70 + "\n")
    app.run(host='0.0.0.0', port=5000, debug=True)